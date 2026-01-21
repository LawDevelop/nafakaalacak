from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)

# Türkçe ay isimleri
TURKISH_MONTHS = {
    1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan',
    5: 'Mayıs', 6: 'Haziran', 7: 'Temmuz', 8: 'Ağustos',
    9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık'
}

def calculate_days_between(start_date, end_date):
    """İki tarih arasındaki gün sayısını hesaplar"""
    delta = end_date - start_date
    return max(0, delta.days)

def calculate_interest(principal, annual_rate, days, day_basis):
    """Basit faiz hesaplar
    
    Args:
        principal: Anapara (TL)
        annual_rate: Yıllık faiz oranı (yüzde olarak, örn: 24 = %24)
        days: Faiz işleyecek gün sayısı
        day_basis: Faize esas gün (360 veya 365)
    
    Returns:
        Hesaplanan faiz tutarı
    """
    if days <= 0:
        return 0
    interest = principal * (annual_rate / 100) * (days / day_basis)
    return round(interest, 2)

def generate_periods(start_date, end_date, monthly_amount, due_day, annual_increase_rate=0, increase_month=1):
    """Nafaka dönemlerini otomatik oluşturur
    
    Args:
        start_date: Başlangıç tarihi
        end_date: Bitiş tarihi
        monthly_amount: Aylık nafaka tutarı
        due_day: Muacceliyet günü (ayın kaçı)
        annual_increase_rate: Yıllık artış oranı (yüzde)
        increase_month: Artışın uygulanacağı ay (1-12)
    
    Returns:
        Dönem listesi
    """
    periods = []
    current_date = start_date
    current_amount = monthly_amount
    current_year = start_date.year
    
    while current_date <= end_date:
        # Yıllık artış kontrolü
        if annual_increase_rate > 0 and current_date.year > current_year:
            if current_date.month >= increase_month:
                years_passed = current_date.year - start_date.year
                current_amount = monthly_amount * ((1 + annual_increase_rate / 100) ** years_passed)
                current_amount = round(current_amount, 2)
        
        # Muacceliyet tarihini belirle
        try:
            due_date = datetime(current_date.year, current_date.month, due_day)
        except ValueError:
            # Ay için geçersiz gün (örn: 31 Şubat), ayın son gününü kullan
            if current_date.month == 12:
                next_month = datetime(current_date.year + 1, 1, 1)
            else:
                next_month = datetime(current_date.year, current_date.month + 1, 1)
            due_date = next_month - timedelta(days=1)
        
        if due_date <= end_date:
            month_name = TURKISH_MONTHS[due_date.month]
            year = due_date.year
            periods.append({
                'id': len(periods) + 1,
                'due_date': due_date.strftime('%Y-%m-%d'),
                'amount': current_amount,
                'description': f"{month_name} {year} Nafakası",
                'interest_rate': None  # Her dönem için özel faiz oranı
            })
        
        # Sonraki aya geç
        current_date = current_date + relativedelta(months=1)
    
    return periods

def calculate_all_interests(periods, calculation_end_date, default_interest_rate, day_basis):
    """Tüm dönemler için faiz hesaplar
    
    Args:
        periods: Dönem listesi
        calculation_end_date: Faiz hesaplama bitiş tarihi
        default_interest_rate: Varsayılan yıllık faiz oranı (yüzde)
        day_basis: Faize esas gün (360 veya 365)
    
    Returns:
        Hesaplama sonuçları
    """
    results = []
    total_principal = 0
    total_interest = 0
    
    for period in periods:
        due_date = datetime.strptime(period['due_date'], '%Y-%m-%d')
        amount = float(period['amount'])
        
        # Dönem için özel faiz oranı varsa onu kullan, yoksa varsayılanı kullan
        period_interest_rate = period.get('interest_rate', default_interest_rate)
        if period_interest_rate is None:
            period_interest_rate = default_interest_rate
        
        days = calculate_days_between(due_date, calculation_end_date)
        interest = calculate_interest(amount, period_interest_rate, days, day_basis)
        
        results.append({
            'id': period['id'],
            'due_date': period['due_date'],
            'due_date_formatted': due_date.strftime('%d.%m.%Y'),
            'description': period.get('description', ''),
            'principal': amount,
            'days': days,
            'interest_rate': period_interest_rate,
            'interest': interest,
            'total': round(amount + interest, 2)
        })
        
        total_principal += amount
        total_interest += interest
    
    return {
        'details': results,
        'summary': {
            'total_principal': round(total_principal, 2),
            'total_interest': round(total_interest, 2),
            'grand_total': round(total_principal + total_interest, 2),
            'period_count': len(results),
            'calculation_date': calculation_end_date.strftime('%d.%m.%Y'),
            'default_interest_rate': default_interest_rate,
            'day_basis': day_basis
        }
    }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate_periods', methods=['POST'])
def api_generate_periods():
    try:
        data = request.json
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')
        monthly_amount = float(data['monthly_amount'])
        due_day = int(data['due_day'])
        annual_increase_rate = float(data.get('annual_increase_rate', 0))
        increase_month = int(data.get('increase_month', 1))
        
        periods = generate_periods(
            start_date, end_date, monthly_amount, due_day,
            annual_increase_rate, increase_month
        )
        
        return jsonify({'success': True, 'periods': periods})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/calculate', methods=['POST'])
def api_calculate():
    try:
        data = request.json
        periods = data['periods']
        calculation_end_date = datetime.strptime(data['calculation_end_date'], '%Y-%m-%d')
        default_interest_rate = float(data['interest_rate'])
        day_basis = int(data['day_basis'])
        
        results = calculate_all_interests(periods, calculation_end_date, default_interest_rate, day_basis)
        
        return jsonify({'success': True, 'results': results})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/validate_calendar', methods=['POST'])
def validate_calendar():
    """Takvim doğrulama - belirtilen tarih aralığındaki tüm günleri kontrol eder"""
    try:
        data = request.json
        start_year = int(data.get('start_year', 2000))
        end_year = int(data.get('end_year', 2030))
        
        validation_results = []
        total_days = 0
        
        for year in range(start_year, end_year + 1):
            year_days = 366 if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0) else 365
            
            # Her ayın gün sayısını kontrol et
            months_info = []
            for month in range(1, 13):
                if month == 2:
                    # Şubat ayı - artık yıl kontrolü
                    days_in_month = 29 if year_days == 366 else 28
                elif month in [4, 6, 9, 11]:
                    days_in_month = 30
                else:
                    days_in_month = 31
                
                # İlk ve son günleri kontrol et
                first_day = datetime(year, month, 1)
                try:
                    last_day = datetime(year, month, days_in_month)
                    months_info.append({
                        'month': TURKISH_MONTHS[month],
                        'days': days_in_month,
                        'first_day': first_day.strftime('%d.%m.%Y'),
                        'last_day': last_day.strftime('%d.%m.%Y')
                    })
                    total_days += days_in_month
                except ValueError as e:
                    return jsonify({'success': False, 'error': f'Takvim hatası: {year} yılı {month}. ay için {str(e)}'})
            
            validation_results.append({
                'year': year,
                'total_days': year_days,
                'is_leap_year': year_days == 366,
                'months': months_info
            })
        
        return jsonify({
            'success': True,
            'validation_results': validation_results,
            'total_years': end_year - start_year + 1,
            'total_days': total_days,
            'message': f'{start_year}-{end_year} yılları arası takvim başarıyla doğrulandı!'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    """Gerçek Excel dosyası oluşturma (openpyxl ile)"""
    try:
        data = request.json
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Nafaka Hesaplama"
        
        # Styles
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="2563EB")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "NAFAKA GEÇMİŞ FAİZ HESAPLAMA RAPORU"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Date
        ws['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:H2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Empty row
        ws.row_dimensions[3].height = 5
        
        # Summary section
        ws['A4'] = "ÖZET BİLGİLER"
        ws['A4'].font = Font(bold=True, size=11)
        ws.merge_cells('A4:B4')
        ws['A4'].fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        summary_data = [
            ['Toplam Anapara:', data['summary']['total_principal']],
            ['Toplam Faiz:', data['summary']['total_interest']],
            ['Genel Toplam:', data['summary']['grand_total']]
        ]
        
        for idx, (label, value) in enumerate(summary_data, start=5):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            ws[f'A{idx}'].font = Font(bold=True)
            ws[f'A{idx}'].border = border
            ws[f'B{idx}'].border = border
            ws[f'B{idx}'].alignment = Alignment(horizontal='right')
        
        # Empty row
        ws.row_dimensions[8].height = 10
        
        # Details header
        headers = ['Sıra', 'Muacceliyet Tarihi', 'Açıklama', 'Anapara (₺)', 'Gün Sayısı', 'Faiz Oranı', 'Faiz (₺)', 'Toplam (₺)']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Details data
        for row_idx, item in enumerate(data['details'], start=10):
            ws.cell(row=row_idx, column=1, value=item['sira'])
            ws.cell(row=row_idx, column=2, value=item['tarih'])
            ws.cell(row=row_idx, column=3, value=item['aciklama'])
            ws.cell(row=row_idx, column=4, value=item['anapara'])
            ws.cell(row=row_idx, column=5, value=item['gun'])
            ws.cell(row=row_idx, column=6, value=item['faiz_orani'])
            ws.cell(row=row_idx, column=7, value=item['faiz'])
            ws.cell(row=row_idx, column=8, value=item['toplam'])
            
            # Apply borders and alignment
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                if col in [1, 5]:  # Sıra ve Gün Sayısı
                    cell.alignment = Alignment(horizontal='center')
                elif col in [4, 7, 8]:  # Para sütunları
                    cell.alignment = Alignment(horizontal='right')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'nafaka_hesaplama_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    # Geliştirme için debug=True, production için debug=False kullanın
    app.run(debug=True, port=5000)
